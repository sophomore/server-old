import os

from openpyxl import load_workbook

from models import OrderMenu, Order, Menu
from mydb import db_session as db

g_menus = {}

def get_menus():
    global g_menus
    if len(g_menus)==0:
        g_menus = {}
        menu = Menu.query.all()
        for m in menu:
            g_menus[m.id] = m
    return g_menus

def print_statement(ordermenus,time):
    time1 = time.strftime('%Y-%m-%d %H:%M:%S')
    menus = get_menus()
    ct = {}
    order = {}
    curry = {}
    twice = {}
    takeout = {}
    t_curry = {}
    t_twice = {}
    t_ct = {}
    for ordermenu in ordermenus:
        name = menus[ordermenu.menu_id].name
        if not ordermenu.takeout:
            if name in order:
                order[name] += 1
            else:
                order[name] = 1
                curry[name] = 0
                twice[name] = 0
                ct[name] = 0
            if ordermenu.curry and ordermenu.twice:
                ct[name] +=1
            elif ordermenu.twice:
                twice[name] +=1
            elif ordermenu.curry:
                curry[name] +=1
        else:
            if name in takeout:
                takeout[name] += 1
            else:
                takeout[name] = 1
                t_curry[name] = 0
                t_twice[name] = 0
                t_ct[name] = 0
            if ordermenu.curry and ordermenu.twice:
                t_ct[name] +=1
            elif ordermenu.twice:
                t_twice[name] +=1
            elif ordermenu.curry:
                t_curry[name] +=1
    string =u'\x1B\x44\x12\x00'
    for key in order:
        if order[key]-curry[key]-twice[key]+ct[key] >0:
            string += u'\x1d\x21\x11'+key+'\x09\x09'+str(order[key]-curry[key]-twice[key]+ct[key])+'\n\x1d\x21\x00'
            string += u'  ㄴ일반\n\n'
        if ct[key]>0:
            string += u'\x1d\x21\x11'+key+'\x09\x09'+str(ct[key])+'\n\x1d\x21\x00'
            string += u'  ㄴ카레\n\n'
            string += u'  ㄴ  곱\n\n'
        if curry[key]>0:
            string += u'\x1d\x21\x11'+key+'\x09\x09'+str(curry[key])+'\n\x1d\x21\x00'
            string += u'  ㄴ카레\n\n'
        if twice[key]>0:
            string += u'\x1d\x21\x11'+key+'\x09\x09'+str(twice[key])+'\n\x1d\x21\x00'
            string += u'  ㄴ  곱\n\n'
    if not  len(takeout) == 0:
        string +=u'------------------포 장------------------\n'
    for key in takeout: 
        if takeout[key]+t_ct[key]-t_curry[key]-t_twice[key] >0:
            string += u'\x1d\x21\x11'+key+'\x09\x09'+str(takeout[key]-t_curry[key]-t_twice[key]+t_ct[key])+'\n\x1d\x21\x00'
            string += u'  ㄴ일반\n\n'
        if t_ct[key]>0:
            string += u'\x1d\x21\x11'+key+'\x09\x09'+str(t_ct[key])+'\n\x1d\x21\x00'
            string += u'  ㄴ카레\n\n'
            string += u'  ㄴ  곱\n\n'
        if t_curry[key]>0:
            string += u'\x1d\x21\x11'+key+'\x09\x09'+str(t_curry[key])+'\n\x1d\x21\x00'
            string += u'  ㄴ카레\n\n'
        if t_twice[key]>0:
            string += u'\x1d\x21\x01'+key+'\x09\x09'+str(t_curry[key])+'\n\x1d\x21\x00'
            string += u'  ㄴ  곱\n\n'
    outstring = u'\x1B\x44\x12\x00'
    outstring =u'                                           \n'
    outstring =u'                                           \n'
    outstring +=u'================전     표================\n\n'
    outstring +=u'주문:'+time1+'\n'
    outstring +=u'----------------------------------------\n'
    outstring +=u'메    뉴\x09\x09    수량\n'
    outstring +=u'----------------------------------------\n'
    outstring +=u''+string
    outstring +=u'----------------------------------------\n\n\n\n'
    outstring +=u'                                        \n'
    outstring +=u'                                        \n'
    outstring +=u'                                        \n'
    outstring += u'\x1bm'
    f2 = open('./statement','w+',encoding="euc-kr")
    string = u'                                            \n'
    print(string,file = f2)
    # print(outstring)
    print(outstring,file = f2)
    f2.close()
    os.system('lpr -P RECEIPT_PRINTER statement')

def print_receipt(orders):
    time = orders.time.strftime('%Y-%m-%d %H:%M:%S')
    menus = get_menus()
    order = {}
    serv = {}
    curry = 0
    twice = 0
    takeout = 0
    summ = orders.totalprice
    for ordermenu in orders.ordermenus:
        name = menus[ordermenu.menu_id]
        if ordermenu.pay != 3:            
            if name in order:
                order[name] += 1
            else:
                order[name] = 1
            if ordermenu.pay != 3:
                if ordermenu.curry:
                    curry+=1
                if ordermenu.twice:
                    twice +=1
                if ordermenu.takeout:
                    takeout +=1
        else:
            if name in serv:
                serv[name] += ordermenu.totalprice
            else:
                serv[name] = ordermenu.totalprice

    orderstring = u'\x1b\x44\x10\x19\x20\x00'
    ser = 0;
    for o in order:
        menuname,a,b = o.name.partition("(수정")
        orderstring +=u'  '+menuname[:9]+'\x09  '+str(order[o])+'\x09'+str(o.price)+'\x09'+str(order[o]*o.price)+'\n'
    for o in serv:
        ser += serv[o]
    if curry>0:
        orderstring +=u'  카레추가\x09  '+str(curry)+'\x092500\x09'+str(2500*curry)+'\n'
    if twice>0:
        orderstring +=u'  곱배기\x09  '+str(twice)+'\x092500\x09'+str(2500*twice)+'\n'
    if takeout>0:
        orderstring +=u'  포장\x09  '+str(takeout)+'\x09500\x09'+str(500*takeout)+'\n'
    orderstring +=u'-----------------------------------------\n'
    orderstring +=u'\x1b\x61\x02합계 : '+str(summ-ser)+'     \n\x1b\x61\x00'
    output = u'                                          \n'
    output += u'                                          \n'
    output +=u'상 호 명: 송호성 쉐프의 돈까스\n'
    output +=u'등록번호: 134-31-16828\n'
    output +=u'대    표: 송호성\n'
    output +=u'전화번호: 031-480-4595\n'
    output +=u'주    소: 경기 안산시 상록구 사동 1165번지\n\n'
    output +=u'주문:'+time+'\n'
    output +=u'-----------------------------------------\n'
    output +=u'  상 품 명\x09수 량\x09단 가\x09금 액\n'
    output +=u'-----------------------------------------\n'
    output +=u''+orderstring
    output +=u'-----------------------------------------\n'
    output +=u'                                         \n'
    output +=u'                                         \n'
    output +=u'                                         \n\n\n\n'
    output +=u'\x1B\x40\x1bm'
    f1 = open('./receipt','w+',encoding="euc-kr")
    # print(output)
    print(output,file = f1)
    f1.close()
    os.system('lpr -P RECEIPT_PRINTER receipt')



def output():
    os.system("mysqldump --user=song --password=Qoswlfdlsnrn pos > ./backup.sql")

def input_two():
    os.system("mysql --user=song --password=Qoswlfdlsnrn pos < backup.sql")

def input():
    menus = get_menus()
    ms = {}
    for menu in menus:
        ms[menus[menu].name] = menu
    wb = load_workbook(filename='../backup.xlsx',read_only = True)
    ws = wb['입출금관리']
    a = lambda x: x>0
    takeout = lambda x: x=="배달"
    f = open("../menus.txt",'w')
    for row in ws.iter_rows(row_offset=1):
        if str(row[0].value) == "합 계":
            break
        else:
            date = str(row[0].value)+' '+str(row[2].value)+':00'
        totalprice = int(row[6].value)
        order = Order(date,totalprice)
        db.add(order)
        db.commit()
        print(order)
        count_twice = get_sign(str(row[13].value),"곱배기추가")
        count_curry = get_sign(str(row[13].value),"카레추가")
        if int(row[8].value) == 0:
            pay = 2
        elif int(row[9].value) ==0:
            pay = 1
        else:
            pay = 4
        ordermenus = str(row[13].value).split(",")
        for o in ordermenus:
            if o.endswith(")"):
                bef,m,aft = o.partition("(")
                num,m,aft = aft.partition(")")
                for i in range(int(num)):
                    if bef in ms:
                        ordermenu = OrderMenu(menu=menus[ms[bef]],order=order, pay=pay,curry=a(count_curry),
                        twice=a(count_twice),takeout=takeout(str(row[3].value)))
                        db.add(ordermenu)
                    else:
                        f.write(bef)
                    count_twice = count_twice - 1
                    count_curry = count_curry - 1
            else:
                if o in ms:
                    ordermenu = OrderMenu(menu=menus[ms[o]],order=order,pay=pay,curry=a(count_curry),
                    twice=a(count_twice),takeout=takeout(str(row[3].value)))
                    db.add(ordermenu)
                else:
                    print(o)
                    f.write(o)
                count_twice = count_twice - 1
                count_curry = count_curry - 1 
        db.commit()
        f.write('\n')
    f.close()

def get_sign(strg,st):
    bef,m,aft = strg.partition(st)
    if aft != "" and aft.startswith("("):
        bef,m,aft = aft[1:].partition(")")
        count = int(bef)
    elif aft.startswith(","):
        count = 1
    else:
        count = 0
    return count

